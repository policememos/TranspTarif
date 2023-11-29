import openpyxl


LOG_FILE = []
fot_tarif_name = 'tarif.xlsx'
zmat_list_name = 'ZMAT_LIST.xlsx'
zspk_all_name = 'ZSPK_ALL.xlsx'

def save_log():
    with open('log.txt', 'w', encoding='utf-8') as logfil:
        for line in LOG_FILE:
            logfil.write(line)
            logfil.write('\n')

def read_xslx(file):
    csv_data = list()
    xlsx = openpyxl.load_workbook(file)
    sheet = xlsx.active
    data = sheet.rows
    for row in data:
        l = list([el.value for el in row])
        csv_data.append(l)
    return csv_data

def parce_tarifs_naming(arrs):
    flag_names = False
    dicts = []
    for raw in arrs:
        if not flag_names:
            raw = raw[4:]
            codes = []
            for name in raw:
                name = name.split('[')[1].replace(',','.')
                s = name.find(']')
                if s == -1:
                    LOG_FILE.append(f'ошибка парсинга кода тарифа {name}')
                    print(f'ошибка парсинга кода тарифа {name}')
                codes.append(name[:s])
            for n in codes:
                n = n.upper()
                tar_name = n
                n = n.split('-')
                tmp = n[1][:2]
                if tmp == 'PO':
                    tmp = n[1]
                elif tmp == 'РО':
                    tmp = n[1]
                elif tmp == 'PО':
                    tmp = n[1]
                elif tmp == 'РO':
                    tmp = n[1]
                else:
                    tmp = 'PO'+n[1]
                if len(n) == 5:
                    dic = {
                        'z': n[0],
                        'fot': tmp,
                        'type': (n[2],),
                        'tonns': n[3].replace('Т','').replace('T',''),
                        'len': n[4].replace('M','').replace('М',''),
                        'tar_name': tar_name,
                        'prices': dict()
                    }
                    dicts.append(dic)
                elif len(n) == 6:
                    dic = {
                        'z': n[0],
                        'fot': tmp,
                        'type': (n[2],n[3]),
                        'tonns': n[4].replace('Т','').replace('T',''),
                        'len': n[5].replace('M','').replace('М',''),
                        'tar_name': tar_name,
                        'prices': dict()
                    }
                    dicts.append(dic)
                else:
                    LOG_FILE.append(f'Аномалия кода тарифа {n}')
                    print(f'Аномалия кода тарифа {n}')
            flag_names = True
        else:
            counter = 4
            for _dict in dicts:
                x = raw[counter]
                if x is None:
                    counter += 1
                    continue
                _dict['prices'].setdefault(raw[0], dict())
                pr = x
                _dict['prices'][raw[0]].setdefault(raw[2], pr)
                counter += 1

            
    return dicts

def check_type(_type):
    if 'CL' in _type:
        if 'ST' in _type:
            return 'CLST'
    if 'CL' in _type:
        return 'CL'
    if 'ST' in _type:
        return 'ST'
    else:
        return 'GT'

def parce_zmat_list(zmat):
    data = []
    for zmat in zmat_list[1:]:
        _type, _tonns, _len, _num, _name = ['GT'], float(zmat[6]), float(zmat[5]), zmat[1], zmat[2]
        if zmat[3] is not None:
            zmat[3] = 'X'
        if zmat[4] is not None:
            zmat[4] = 'X'
        if zmat[3]== 'X' and zmat[4] == 'X':
            _type = ['CL', 'ST']
        elif zmat[3] == 'X':
            _type = ['ST']
        elif zmat[4] == 'X':
            _type = ['CL']
        data.append({
            'type': _type,
            'tonns': _tonns,
            'len': _len,
            'num': _num,
            'naming': _name
        })
    return data

def zm_finder(tarif, zmat_data, type_cheker):
    fitting_zmats = []
    tar_tonns = float(tarif['tonns'])
    tar_len = float(tarif['len'])
    if type_cheker == 'CLST':
        for zm in zmat_data:
            if 'CL' in zm['type']:
                if 'ST' in zm['type']:
                    fitting_zmats.append(zm)
    if type_cheker == 'CL':
        for zm in zmat_data:
            if len(zm['type']) == 1:
                if 'CL' in zm['type']:
                    fitting_zmats.append(zm)
    if type_cheker == 'ST':
        for zm in zmat_data:
            if len(zm['type']) == 1:
                if 'ST' in zm['type']:
                    fitting_zmats.append(zm)
    if type_cheker == 'GT':
        for zm in zmat_data:
            if len(zm['type']) == 1:
                if 'GT' in zm['type']:
                    fitting_zmats.append(zm)

    sorted_dict = sorted(fitting_zmats, key = lambda x: (x['tonns'], x['len']))

    LOG_FILE.append(f'\n{tarif["tar_name"]}\nВид: {type_cheker}\nВес: {tar_tonns}\nДлина: {tar_len}')
    print(f'\n{tarif["tar_name"]}\nВид: {type_cheker}\nВес: {tar_tonns}\nДлина: {tar_len}')
    find_fl = False
    for zm in sorted_dict:
        # проверка аномального тарифа
        if tar_tonns == 22.0 and tar_len == 13.5:
            if zm['tonns'] == 22.0 and zm['len'] == 13.5:
                LOG_FILE.append(f'Определён тариф {zm["type"]}{zm["tonns"]}T{zm["len"]}M---{zm["num"]} {zm["naming"]}')
                print(f'Определён тариф {zm["type"]}{zm["tonns"]}T{zm["len"]}M---{zm["num"]} {zm["naming"]}')
                find_fl = True
                return zm['num'], zm['len']

        if tar_len in (13.0, 13.5):
            tar_len = 12.0

        if tar_tonns <= zm['tonns']:
            if tar_len <= zm['len']:
                LOG_FILE.append(f'Определён тариф {zm["type"]}-{zm["tonns"]}T-{zm["len"]}M---{zm["num"]} {zm["naming"]}')
                print(f'Определён тариф {zm["type"]}-{zm["tonns"]}T-{zm["len"]}M---{zm["num"]} {zm["naming"]}')
                find_fl = True
                return zm['num'], zm['len']
    if not find_fl:
        return find_fl, False

def parce_mat_names(tarifs, zmat, skipping):
    tarif_with_number = []
    zmat_data = parce_zmat_list(zmat)
    for tarif in tarifs:
        if skipping:
            if any(lambda x: 1 for x in ['SM','PS','LG'] if x in tarif['type']):
                LOG_FILE.append(f"\nСкипаем тариф {tarif['z']}-{tarif['fot']}-{''.join(tarif['type'])}-{tarif['tonns']}T-{tarif['len']}M\nОн Врф/Пцс/Логистика\nПараметр skipping=1\n")
                print(f"\nСкипаем тариф {tarif['z']}-{tarif['fot']}-{''.join(tarif['type'])}-{tarif['tonns']}T-{tarif['len']}M\nОн Врф/Пцс/Логистика\nПараметр skipping=1\n")
                continue
            
        tr_type_cheker = check_type(tarif['type'])
        mat_num, mat_len = zm_finder(tarif, zmat_data, tr_type_cheker)
        if not mat_num:
            LOG_FILE.append('Не нашлось материала автодоставки')
            print('Не нашлось материала автодоставки')
            continue

        tarif.setdefault('numeber', mat_num)
        try:
            int_mat_len = int(mat_len)
        except:
            int_mat_len = mat_len
        tarif['len'] = int_mat_len
        tarif_with_number.append(tarif)
    return tarif_with_number

def fill_first_sheet(sheet_my, last_row):
    line_flag = False
    if not line_flag:
        line_counter = int(last_row)+1
        line_flag = True
    for tarif in mapped_tarifs:
        a_col = ('ZR11', 'ZW91')
        b_col = ('73', '*')
        f_col = tarif['fot']
        g_col = str(tarif['tonns'])+'Т'
        h_col = str(tarif['len'])+'М'
        i_col = 'X' if 'ST' in tarif['type'] else ''
        j_col = 'X' if 'CL' in tarif['type'] else ''
        if not tarif.get('numeber', False):
            ttp = "-".join(tarif['type'])
            LOG_FILE.append(f'\nОшибка при создании новго экселя с тарифом Z-{f_col}-{ttp}-{g_col}-{h_col}\n---> Нет подходящего материала в zmat_list\n')
            print(f'\nОшибка при создании новго экселя с тарифом Z-{f_col}-{ttp}-{g_col}-{h_col}\n---> Нет подходящего материала в zmat_list\n')
            continue
        c_col = tarif['numeber']
        for start_mpl, end_mpl in tarif['prices'].items():
            if isinstance(end_mpl, dict):
                for _end_mpl, _pr in end_mpl.items():
                    d_col = start_mpl
                    e_col = _end_mpl
                    k_col = _pr
                    for i in a_col:
                        sheet_my.cell(row=line_counter, column=1, value=i)
                        if i == 'ZR11':
                            sheet_my.cell(row=line_counter, column=2, value=b_col[0])
                        else:
                            sheet_my.cell(row=line_counter, column=2, value=b_col[1])
                        sheet_my.cell(row=line_counter, column=3, value=c_col)
                        sheet_my.cell(row=line_counter, column=4, value=d_col)
                        sheet_my.cell(row=line_counter, column=5, value=e_col)
                        sheet_my.cell(row=line_counter, column=6, value=f_col)
                        sheet_my.cell(row=line_counter, column=7, value=g_col)
                        sheet_my.cell(row=line_counter, column=8, value=h_col)
                        sheet_my.cell(row=line_counter, column=9, value=i_col)
                        sheet_my.cell(row=line_counter, column=10, value=j_col)
                        sheet_my.cell(row=line_counter, column=11, value=k_col)
                        line_counter += 1

def backup_fill_first_sheet(sheet_my):
    ed_sheet = sheet_my
    for tarif in mapped_tarifs:
        line_counter = int(last_row)+1
        a_col = ('ZR11', 'ZW91')
        b_col = ('73', '*')
        f_col = tarif['fot']
        g_col = tarif['tonns']+'T'
        h_col = tarif['len']+'M'
        i_col = 'X' if 'ST' in tarif['type'] else ''
        j_col = 'X' if 'CL' in tarif['type'] else ''
        if not tarif.get('numeber', False):
            ttp = "-".join(tarif['type'])
            LOG_FILE.append(f'Ошибка при создании новго экселя с тарифом Z-{f_col}-{ttp}-{g_col}-{h_col}')
            print(f'Ошибка при создании новго экселя с тарифом Z-{f_col}-{ttp}-{g_col}-{h_col}')
            continue
        c_col = tarif['numeber']
        for start_mpl, end_mpl in tarif['prices'].items():
            if isinstance(end_mpl, dict):
                for _end_mpl, _pr in end_mpl.items():
                    d_col = start_mpl
                    e_col = _end_mpl
                    k_col = _pr
                    for i in a_col:
                        ed_sheet.cell(row=line_counter, column=1, value=i)
                        if i == 'ZR11':
                            ed_sheet.cell(row=line_counter, column=2, value=b_col[0])
                        else:
                            ed_sheet.cell(row=line_counter, column=2, value=b_col[1])
                        ed_sheet.cell(row=line_counter, column=3, value=c_col)
                        ed_sheet.cell(row=line_counter, column=4, value=d_col)
                        ed_sheet.cell(row=line_counter, column=5, value=e_col)
                        ed_sheet.cell(row=line_counter, column=6, value=f_col)
                        ed_sheet.cell(row=line_counter, column=7, value=g_col)
                        ed_sheet.cell(row=line_counter, column=8, value=h_col)
                        ed_sheet.cell(row=line_counter, column=9, value=i_col)
                        ed_sheet.cell(row=line_counter, column=10, value=j_col)
                        ed_sheet.cell(row=line_counter, column=11, value=k_col)
                        line_counter += 1
    return ed_sheet

zmat_list = read_xslx(zmat_list_name)
fot_tarif = read_xslx(fot_tarif_name)
tarif_codes = parce_tarifs_naming(fot_tarif)
mapped_tarifs = parce_mat_names(tarif_codes, zmat_list, skipping=True)
zspk_all_xlsx = openpyxl.load_workbook(zspk_all_name)
first_sheet = zspk_all_xlsx.active
last_row = first_sheet.dimensions.split(':')[1][1:]
fill_first_sheet(first_sheet, last_row)
zspk_all_xlsx.save('result.xlsx')
save_log()